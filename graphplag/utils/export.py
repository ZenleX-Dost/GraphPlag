"""
Export utilities for generating PDF and Excel reports from plagiarism detection results.
"""
from typing import List, Optional
from pathlib import Path
from datetime import datetime
import io

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from graphplag.core.models import PlagiarismReport, PlagiarismMatch


class PDFReportGenerator:
    """Generate PDF reports with highlighted plagiarism matches."""
    
    def __init__(self):
        """Initialize PDF generator."""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Highlight style for plagiarized text
        self.styles.add(ParagraphStyle(
            name='Highlighted',
            parent=self.styles['Normal'],
            backColor=colors.HexColor('#FFFF00'),
            textColor=colors.HexColor('#000000')
        ))
        
        # Match style
        self.styles.add(ParagraphStyle(
            name='Match',
            parent=self.styles['Normal'],
            backColor=colors.HexColor('#FFE6E6'),
            leftIndent=20,
            rightIndent=20
        ))
    
    def generate_report(
        self,
        report: PlagiarismReport,
        output_path: str,
        include_full_text: bool = True
    ) -> str:
        """
        Generate PDF report from plagiarism detection results.
        
        Args:
            report: PlagiarismReport object
            output_path: Path to save PDF file
            include_full_text: Include full document texts
            
        Returns:
            Path to generated PDF
        """
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        story = []
        
        # Title
        story.append(Paragraph("Plagiarism Detection Report", self.styles['CustomTitle']))
        story.append(Spacer(1, 12))
        
        # Summary information
        summary_data = [
            ["Document 1:", report.document1.doc_id or "N/A"],
            ["Document 2:", report.document2.doc_id or "N/A"],
            ["Overall Similarity:", f"{report.similarity_score:.2%}"],
            ["Threshold:", f"{report.threshold:.2%}"],
            ["Result:", "PLAGIARISM DETECTED" if report.is_plagiarism else "NO PLAGIARISM"],
            ["Method:", report.method.upper()],
            ["Matches Found:", str(len(report.matches))],
            ["Processing Time:", f"{report.processing_time:.3f}s"],
            ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8E8E8')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Kernel scores if available
        if report.kernel_scores:
            story.append(Paragraph("Kernel Scores:", self.styles['Heading2']))
            kernel_data = [[k.upper(), f"{v:.4f}"] for k, v in report.kernel_scores.items()]
            kernel_table = Table(kernel_data, colWidths=[2*inch, 2*inch])
            kernel_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F0F0F0')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            story.append(kernel_table)
            story.append(Spacer(1, 20))
        
        # Matches
        if report.matches:
            story.append(Paragraph(f"Detected Matches ({len(report.matches)}):", self.styles['Heading2']))
            story.append(Spacer(1, 12))
            
            for idx, match in enumerate(report.matches[:20], 1):  # Limit to 20 matches
                story.append(Paragraph(f"<b>Match {idx}</b> (Similarity: {match.similarity:.2%})", self.styles['Heading3']))
                
                # Document 1 text
                story.append(Paragraph("<b>Document 1:</b>", self.styles['Normal']))
                story.append(Paragraph(match.text1, self.styles['Match']))
                story.append(Spacer(1, 6))
                
                # Document 2 text
                story.append(Paragraph("<b>Document 2:</b>", self.styles['Normal']))
                story.append(Paragraph(match.text2, self.styles['Match']))
                story.append(Spacer(1, 12))
                
                if idx < len(report.matches):
                    story.append(Spacer(1, 6))
        
        # Full texts (optional)
        if include_full_text:
            story.append(PageBreak())
            story.append(Paragraph("Full Document Texts", self.styles['Heading2']))
            story.append(Spacer(1, 12))
            
            story.append(Paragraph("<b>Document 1:</b>", self.styles['Heading3']))
            doc1_text = " ".join([s.text for s in report.document1.sentences])
            story.append(Paragraph(doc1_text[:5000], self.styles['Normal']))  # Limit length
            story.append(Spacer(1, 20))
            
            story.append(Paragraph("<b>Document 2:</b>", self.styles['Heading3']))
            doc2_text = " ".join([s.text for s in report.document2.sentences])
            story.append(Paragraph(doc2_text[:5000], self.styles['Normal']))  # Limit length
        
        # Build PDF
        doc.build(story)
        
        return output_path
    
    def generate_to_bytes(self, report: PlagiarismReport) -> bytes:
        """Generate PDF report to bytes (for API responses)."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        
        # Similar to generate_report but builds to buffer
        story = []
        story.append(Paragraph("Plagiarism Detection Report", self.styles['CustomTitle']))
        # ... (simplified version)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.read()


class ExcelReportGenerator:
    """Generate Excel reports with color-coded plagiarism matches."""
    
    def __init__(self):
        """Initialize Excel generator."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def generate_report(
        self,
        report: PlagiarismReport,
        output_path: str
    ) -> str:
        """
        Generate Excel report from plagiarism detection results.
        
        Args:
            report: PlagiarismReport object
            output_path: Path to save Excel file
            
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        self._create_summary_sheet(wb, report)
        
        # Matches sheet
        if report.matches:
            self._create_matches_sheet(wb, report)
        
        # Sentences sheet
        self._create_sentences_sheet(wb, report)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(self, wb: Workbook, report: PlagiarismReport):
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "Plagiarism Detection Report"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # Summary data
        row = 3
        summary_data = [
            ("Document 1:", report.document1.doc_id or "N/A"),
            ("Document 2:", report.document2.doc_id or "N/A"),
            ("Overall Similarity:", f"{report.similarity_score:.2%}"),
            ("Threshold:", f"{report.threshold:.2%}"),
            ("Result:", "PLAGIARISM DETECTED" if report.is_plagiarism else "NO PLAGIARISM"),
            ("Method:", report.method.upper()),
            ("Matches Found:", len(report.matches)),
            ("Processing Time:", f"{report.processing_time:.3f}s"),
            ("Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Result highlighting
        result_row = 7  # "Result:" row
        if report.is_plagiarism:
            ws[f'B{result_row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            ws[f'B{result_row}'].font = Font(bold=True, color="CC0000")
        else:
            ws[f'B{result_row}'].fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            ws[f'B{result_row}'].font = Font(bold=True, color="00CC00")
        
        # Kernel scores
        if report.kernel_scores:
            row += 2
            ws[f'A{row}'] = "Kernel Scores:"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            for kernel, score in report.kernel_scores.items():
                ws[f'A{row}'] = kernel.upper()
                ws[f'B{row}'] = f"{score:.4f}"
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_matches_sheet(self, wb: Workbook, report: PlagiarismReport):
        """Create matches sheet with highlighting."""
        ws = wb.create_sheet("Matches")
        
        # Headers
        headers = ["Match #", "Similarity", "Doc1 Text", "Doc2 Text", "Doc1 Position", "Doc2 Position"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for idx, match in enumerate(report.matches, 2):
            ws.cell(idx, 1, idx - 1)
            ws.cell(idx, 2, f"{match.similarity:.2%}")
            ws.cell(idx, 3, match.text1)
            ws.cell(idx, 4, match.text2)
            ws.cell(idx, 5, f"{match.start1}-{match.end1}")
            ws.cell(idx, 6, f"{match.start2}-{match.end2}")
            
            # Color code by similarity
            if match.similarity >= 0.9:
                fill_color = "FFCCCC"  # High similarity - red
            elif match.similarity >= 0.7:
                fill_color = "FFEECC"  # Medium - orange
            else:
                fill_color = "FFFFCC"  # Low - yellow
            
            for col in range(1, 7):
                ws.cell(idx, col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                ws.cell(idx, col).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_sentences_sheet(self, wb: Workbook, report: PlagiarismReport):
        """Create sheet with all sentences."""
        ws = wb.create_sheet("Sentences")
        
        # Headers
        ws['A1'] = "Document"
        ws['B1'] = "Sentence #"
        ws['C1'] = "Text"
        
        for col in range(1, 4):
            ws.cell(1, col).font = Font(bold=True)
            ws.cell(1, col).fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
        
        # Document 1 sentences
        row = 2
        for idx, sent in enumerate(report.document1.sentences, 1):
            ws.cell(row, 1, "Document 1")
            ws.cell(row, 2, idx)
            ws.cell(row, 3, sent.text)
            row += 1
        
        # Document 2 sentences
        for idx, sent in enumerate(report.document2.sentences, 1):
            ws.cell(row, 1, "Document 2")
            ws.cell(row, 2, idx)
            ws.cell(row, 3, sent.text)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 80
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    def generate_to_bytes(self, report: PlagiarismReport) -> bytes:
        """Generate Excel report to bytes (for API responses)."""
        wb = Workbook()
        wb.remove(wb.active)
        self._create_summary_sheet(wb, report)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()
